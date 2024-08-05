from openpyxl.reader.excel import load_workbook
from _shared import osu_apis
from _shared import Team
import ossapi
from ossapi import MatchEventType

api = osu_apis.get_osu_api()

def read_teams():
    teams_workbook_path = '../_shared/_datas/teams.xlsx'
    col_index = {
        'uid' : 0,
        'name' : 1,
        'team' : 3,
        'group': 4
    }

    teams = {}
    wb = load_workbook(filename=teams_workbook_path, read_only=True)
    ws = wb['Teams']
    for row in ws.iter_rows(min_row=2):
        player = Team.Player(
            row[col_index['uid']].value,
            row[col_index['name']].value,
            row[col_index['group']].value
        )
        player_team = row[col_index['team']].value
        if not (player_team in teams):
            teams[player_team] = Team.Team(player_team, [])
        teams[player_team].add_player(player)
    return teams


def read_mappools():
    mappools_workbook_path = '_datas/mappools.xlsx'
    col_index = {
        'pick' : 1,
        'map_id' : 4
    }

    mappools = []
    wb = load_workbook(filename=mappools_workbook_path, read_only=True)
    ws = wb['Mappools']
    for row in ws.iter_rows(min_row=2):
        beatmap = {
            'pick' : row[col_index['pick']].value,
            'map_id' : row[col_index['map_id']].value,
        }
        mappools.append(beatmap)
    return mappools

def read_mplinks():
    mplink_workbook_path = '_datas/mplinks.xlsx'
    col_index = {
        'mplink' : 0
    }

    mplinks = []
    wb = load_workbook(filename=mplink_workbook_path, read_only=True)
    ws = wb['Mplinks']
    for row in ws.iter_rows(min_row=2):
        mplink = row[col_index['mplink']].value
        if type(mplink) is str:
            mplink = mplink.split('/')[-1]
        mplinks.append(mplink)
    return mplinks


def read_match(mplink: str | int):
    '''
    Get all raw datas that played in certain match room.

    :param mplink: match room id. stable match only , for lazer matches, using 'read_match_lazer'

    :return:
        a dict objct contain following three keys. <br>
        name - string, match room title. <br>
        id - int, match room id, equals to mplink. <br>
        games - all played maps infos.
    '''
    result = {
        'name': None,
        'id': None,
        'games': []
    }
    try:
        match = api.match(mplink)
        result['name'] = match.match.name
        result['id'] = match.match.id
        for event in match.events:
            if event.detail.type == MatchEventType.OTHER:
                game = event.game
                result['games'].append(game)
    except ValueError as e:
        print('Invalid mplink : ' + str(mplink))
    return result

if __name__ == '__main__':
    # match = read_match(114888291)
    # print(match)
    # print(read_teams())
    # print(read_mappools())
    # print(read_mplinks())
    # print(read_match(read_mplinks()[0]))
